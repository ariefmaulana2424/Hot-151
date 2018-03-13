import openpyxl
import pprint
import datetime
import calendar
import time
from datetime import timedelta
##import numpy as np

wb = openpyxl.load_workbook('GA+Paket.xlsx')
ws = wb['Sheet1']



def indices( mylist, value):
    return [i for i,x in enumerate(mylist) if x==value]
	

list_row		   = []
list_dompul        = []
list_msisidn_B     = []
list_trx_date      = []
list_package_price = []
list_ir_date	   = []
list_type_paket    = []
list_point         = []


	
for col in ws['A']:
	list_row.append(col.value)

for col in ws['B']:
	list_dompul.append(col.value)
	
for col in ws['C']:
	list_msisidn_B.append(col.value)

for col in ws['D']:
	list_trx_date.append(col.value)

for col in ws['E']:
	list_ir_date.append(col.value)

for col in ws['F']:
	list_package_price.append(col.value)

for col in ws['G']:
	list_type_paket.append(col.value)

for col in ws['H']:
	list_point.append(col.value)
	


list_row_without_header_no = list_row[1:]
max_count = max(list_row_without_header_no)
print (max_count)    

for x in range(1,len(list_msisidn_B)):
	list_row_extract      = list_row[x]
#----------------------------------------------------------validasi 1 check XL Number------------------------------------------------------------------#

	if str(list_msisidn_B[x])[0:5] == '62817' or str(list_msisidn_B[x])[0:5] == '62818' or str(list_msisidn_B[x])[0:5] == '62819' or str(list_msisidn_B[x])[0:5] == '62859' or str(list_msisidn_B[x])[0:5] == '62877' or str(list_msisidn_B[x])[0:5] == '62878':
		print ("ini test case ke : ",x)
		print ("")
		print ("validasi 1 passed")
		print ("")
		validasi1 = 1
		
		
		
#-------------------------------------------------validasi 2 Check 1st transaction of B number--------------------------------------------------------#
		
		
		
		date_based_on_index_list = []
		mylist1 = list_msisidn_B[x] #tampung nilai msisdn berdasarkan index x
		#print ("MSISDN B 	    : ",mylist1)

		index_dompul = indices(list_msisidn_B, mylist1) #cari index array, dari input nomor yang di dapat dari mylist
	
		#print ("Total Date	    : ", list_trx_date)
		#print ("index_dompul	    : ",index_dompul)
	
		for y in range(0,len(index_dompul)):
			date_based_on_index = list_trx_date[index_dompul[y]] # ambil nilai date berdasarkan index dompul 2020,20201
			date_based_on_index_list.append(date_based_on_index) # tambahkan nilai tanggal yang di dapat ke baris 13
		
		date_based_on_index_list_min = min(date_based_on_index_list)
		
		#print ("index list : ", date_based_on_index_list)
		#print ("index min : ",date_based_on_index_list_min)
		#print ("ini date x : ", list_trx_date[x])
		
		if list_trx_date[x] == date_based_on_index_list_min:
			print ("validasi 2 passed")
			print ("")
			validasi2 = 1
			
			
#--------------------------------------------------------Validasi 3 TRX DATE Maximum IR+30----------------------------------------------------#			
			
		
			#print ("ini list trx date :", list_trx_date[x])
			#print ("ini list IR date     :", list_ir_date[x])
			
			list_trx_date_datetime = datetime.datetime.strptime(list_trx_date[x],"%d-%m-%Y %H:%M:%S").date()
			list_ir_date_datetime     = datetime.datetime.strptime(list_ir_date[x],"%d-%m-%Y %H:%M:%S" ).date()
			
			#print ("ini list trx date  datetime   :", list_trx_date_datetime)
			#print ("ini list IR     date  datetime   :", list_ir_date_datetime)
			list_ir_date_datetime += timedelta(days=30)
			
			#print ("ini list ir date  datetime tambah 30 hari   :", list_ir_date_datetime)
			#list_trx_date_plus_30 = tambahbulan (list_trx_date[x])
			#print ("ini list trx date +30 :", list_trx_date_plus_30)
			
			if list_ir_date_datetime >= list_trx_date_datetime:
				print ('Validasi 3 Passed')
				print ("")
				validasi3 = 1
				
				
				
#----------------------------------------------------Validasi 4 trx Amount minimum 25k-----------------------------------------------------#			
				
				if list_package_price[x] >= 25000 and list_type_paket[x] == 'Non Flash Sale':
					print("Validasi 4 Passed")
					print ("")
					validasi4 = 1
					
					
					
#--------------------------------------------------Validasi 5 Point 50K-----------------------------------------------------------#						
					
					if list_point[x] == 50000:
						print ("Validasi 5 Passed")
						
						validasi5 = 1
						
						
						
					#else Validasi 5
					else :
						print ("Validasi 5 Failed")
						validasi5 = 0
					
					
				#else Validasi 4	
				else:
					print("Validasi 4 Failed")
					validasi4 = 0
				

			#else Validasi3
			else:
				print ('Validasi 3 Failed')
				print ("")
				validasi3 = 0
		
		
		#else Validasi 2
		else:
			print ("validasi 2 failed")
			print ("")
			validasi2 = 0
		
	#else validasi 1
	else:
		print ("ini test case ke : ",x)	
		print ("validasi 1 failed")
		print ("")
		validasi1 = 0
		
	if ( validasi1 and validasi2 and validasi3 and validasi4 and validasi5 == 1 ) and x != max_count:
		print("f1")
		message=["Passed"]
		for row, entry in enumerate(message, start=1):
			for x in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
				
	elif ( validasi1 and validasi2 and validasi3 and validasi4 and validasi5 == 1):
		print("f2")
		message=["Passed"]
		for row, entry in enumerate(message, start=1):
			for xx in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
		wb.save('GA+Paket.xlsx')
		
	elif (validasi2 == 0 or validasi3 == 0 or validasi4 == 0 ) and list_point[x] == 0 and x != max_count:
		print("f3")
		message=["Passed"]
		for row, entry in enumerate(message, start=1):
			for xx in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
		
		
	elif (validasi2 == 0 or validasi3 == 0 or validasi4 == 0 ) and list_point[x] == 0 :
		print("f4")
		message=["Passed"]
		for row, entry in enumerate(message, start=1):
			for xx in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
		wb.save('GA+Paket.xlsx')
		
	elif ( validasi1 != 2 and validasi2 != 2 and validasi3  != 2 and validasi4  != 2 and validasi5 != 2) and x != max_count:
		print("f5")
		message=["Failed"]
		for row, entry in enumerate(message, start=1):
			for xx in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
		
	elif ( validasi1 != 2 and validasi2  != 2 and validasi3 != 2 and validasi4 != 2 and validasi5 != 2):
		print("f6")
		message=["Failed"]
		for row, entry in enumerate(message, start=1):
			for xx in range(len(list_row)):
				column_cell = 'I'
				ws[column_cell+str(row+list_row_extract)]=str(entry)
		print ("--------------------------------------------------------------------")
		wb.save('GA+Paket.xlsx')
	
		

